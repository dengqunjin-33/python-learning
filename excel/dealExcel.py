import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy

readWb = xlrd.open_workbook("C:/Users/11469/Desktop/xxx/xxx2.xls")
readWb2 = openpyxl.load_workbook("C:/Users/11469/Desktop/xxx/xxx1.xlsx", data_only=True)
readWb3 = xlrd.open_workbook("C:/Users/11469/Desktop/xxx/test3.xls")
writeWb = xlwt.Workbook()
writeWb = copy(readWb3)
# 设置日期格式
dateStyle = xlwt.XFStyle()
dateStyle.num_format_str = 'yyyy-mm-dd'

# 设置垂直居中
wordStyle = xlwt.XFStyle()
al = xlwt.Alignment()
al.horz = 0x02  # 设置水平居中
al.vert = 0x01  # 设置垂直居中
dateStyle.alignment = al
wordStyle.alignment = al

shNames = readWb.sheet_names()
startSheet = 0
for i in range(len(shNames)):
    rsh = readWb.sheet_by_name(shNames[i])
    if str(6.19) == str(rsh.name):
        continue
    rsh2 = readWb3.sheet_by_name(rsh.name)
    print(len(readWb3.sheet_names()))
    print("读：{}", str(rsh.name))
    print("写：{}", str(rsh2.name))
    print(str(rsh.name) == str(rsh2.name))
    while str(rsh.name) != str(rsh2.name):
        startSheet = startSheet+1
        rsh2 = readWb3.sheet_by_index(startSheet)
    count = 9
    for row in range(7, 31):
        if writeWb.get_sheet(startSheet) is None:
            continue
        wsh = writeWb.get_sheet(startSheet)
        # 喷浆
        pengjiang = rsh.cell_value(row, 11)
        # 总施工
        zongshigong = rsh.cell_value(row, 10)
        # 水泥
        zhuangshu = rsh.cell_value(row, 6)
        # 桩号
        zhuanghao = rsh.cell_value(row, 3)
        # 日期
        date = rsh.cell_value(row, 1)
        if count <= 22:
            print('写入')
            # 日期格式
            wsh.write_merge(count, count, 1, 2, date, dateStyle)
            wsh.write_merge(count, count, 3, 3, zhuanghao, wordStyle)
            wsh.write_merge(count, count, 4, 4, pengjiang, wordStyle)
            wsh.write_merge(count, count, 9, 9, zhuangshu, wordStyle)
            wsh.write_merge(count, count, 13, 13, zongshigong, wordStyle)
            count = count + 1
            print('写入后', count)
        if count > 22:
            print('进来了呦')
            count = 9
            startSheet = startSheet+1
            print('startSheet{}', startSheet)
        print(count)

# sheets = readWb2.sheetnames
# startSheet = 0
# for i in range(len(sheets)):
#     sheet = readWb2.get_sheet_by_name(sheets[i])
#     for index in range(12, 39, 4):
#         zh = sheet.cell(row=index, column=4).value
#         if str(rsh.name) != str(rsh2.name):
#             ++startSheet
#             rsh2 = readWb3.sheet_by_name(shNames[startSheet])
#         wsh = writeWb.get_sheet(startSheet)
#         if zh is not None:
#             flag = True
#             for row in range(9, 33):
#                 zhuanghao = rsh2.cell_value(row, 3)
#                 if zhuanghao == str(zh):
#                     wsh.write_merge(row, row, 11, 11, str(sheet.cell(row=index, column=44).value), wordStyle)
#                     wsh.write_merge(row, row, 12, 12, str(sheet.cell(row=(index + 3), column=45).value), wordStyle)
#                     flag = False
#                     continue
#             if flag:
#                 startSheet += 1

writeWb.save("C:/Users/11469/Desktop/xxx/test.xls")
print("--------------写入完成------------")
